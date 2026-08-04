import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FONT = 'Arial'
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT, italic=True, size=10, color="595959")
BODY_FONT = Font(name=FONT, size=10)
BOLD_BODY = Font(name=FONT, size=10, bold=True)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FLAG_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

wb = Workbook()

def style_header_row(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def write_df(ws, df, start_row, start_col=1, index=True, pct_cols=None, highlight_col=None, highlight_thresh=None):
    pct_cols = pct_cols or []
    r = start_row
    headers = ([df.index.name or ''] if index else []) + list(df.columns)
    for j, h in enumerate(headers):
        ws.cell(row=r, column=start_col+j, value=h)
    style_header_row(ws, r, len(headers))
    r += 1
    for idx, row in df.iterrows():
        c = start_col
        if index:
            ws.cell(row=r, column=c, value=idx).font = BODY_FONT
            ws.cell(row=r, column=c).border = BORDER
            c += 1
        for colname, val in row.items():
            cell = ws.cell(row=r, column=c)
            if isinstance(val, (int, float, np.floating, np.integer)):
                cell.value = float(val) if isinstance(val, (np.floating, float)) else int(val)
                if colname in pct_cols:
                    cell.number_format = '0.00"%"'
                elif isinstance(val, float):
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0'
            else:
                cell.value = val
            cell.font = BODY_FONT
            cell.border = BORDER
            c += 1
        r += 1
    return r  # next free row

def autosize(ws, max_col, min_width=10, max_width=45):
    for c in range(1, max_col+1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for cell in ws[col_letter]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_width, max_len + 3)

def title_block(ws, title, subtitle=None, row=1):
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    r = row+1
    if subtitle:
        ws.cell(row=r, column=1, value=subtitle).font = SUBTITLE_FONT
        r += 1
    return r + 1

# ============================================================
# SHEET 1: Executive Summary
# ============================================================
ws = wb.active
ws.title = "Executive Summary"
r = title_block(ws, "Credit EDA Case Study — Executive Summary",
                 "Identifying applicant patterns linked to payment difficulty (TARGET=1) | Home Credit Default Risk dataset")

summary_points = [
    ("Dataset overview",
     "application_data.csv: 307,511 loan applications x 122 fields. previous_application.csv: 1,670,214 prior loan records x 37 fields, linked via SK_ID_CURR."),
    ("Class imbalance",
     "TARGET is heavily imbalanced: 91.93% of applicants repaid on time (TARGET=0) vs 8.07% had payment difficulty (TARGET=1). Any comparison between segments must account for this skew — raw counts will always favor TARGET=0."),
    ("Data quality",
     "67 of 122 columns had missing values; 49 columns exceeded 40% missing (mostly building/apartment detail fields) and were dropped for the core analysis. DAYS_EMPLOYED contained a placeholder anomaly (365243, ~18% of rows) representing pensioners/unemployed, corrected to NULL before deriving YEARS_EMPLOYED."),
    ("Strongest risk indicators found",
     "1) EXT_SOURCE_2 and EXT_SOURCE_3 (external credit bureau scores) — sharply lower among defaulters (median 0.44/0.38) vs non-defaulters (0.57/0.55). "
     "2) Applicant age — default rate falls monotonically from 11.4% (20-30 yrs) to 4.9% (60-70 yrs). "
     "3) Years employed — defaulters have shorter tenure (median 3.4 yrs vs 4.6 yrs). "
     "4) Prior loan refusals — clients with 5+ prior refusals from this lender default at 14.1% vs 7.0% for clients with zero prior refusals — a near-monotonic relationship. "
     "5) Income type / occupation — Low-skill Laborers (17.2%), Drivers (11.3%) and Working income type (9.6%) show materially higher default rates than Pensioners (5.4%) or State servants (5.8%). "
     "6) Education — Lower secondary education defaults at 10.9% vs 5.4% for Higher education. "
     "7) Housing — Rented apartment (12.3%) and Living with parents (11.7%) default more than House/apartment owners (7.8%)."),
    ("Weak / non-indicators",
     "Loan amount, credit-to-income ratio, and annuity-to-income ratio show only marginal differences between the two groups at the median — the amount requested is a weaker signal than the applicant's underlying financial stability and history."),
    ("Correlation structure shifts",
     "The correlation between AMT_INCOME_TOTAL and AMT_CREDIT/AMT_ANNUITY collapses from ~0.35-0.42 (TARGET=0) to ~0.04-0.05 (TARGET=1) — among defaulters, loan size is no longer well-anchored to income, a useful underwriting flag."),
    ("Recommended actions",
     "Weight EXT_SOURCE scores and applicant age heavily in scoring; treat repeated prior refusals as a strong escalation trigger (manual review / higher rate tier); apply extra scrutiny to Low-skill Laborers, Drivers and short-tenure (<2 yr) applicants; do not rely on credit-to-income ratio alone as a risk cut-off."),
]

for label, text in summary_points:
    ws.cell(row=r, column=1, value=label).font = BOLD_BODY
    ws.cell(row=r, column=1).alignment = Alignment(vertical='top')
    cell = ws.cell(row=r, column=2, value=text)
    cell.font = BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = max(30, 15*(len(text)//90 + 1))
    r += 2

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 110

# ============================================================
# SHEET 2: Data Quality (missing values, outliers)
# ============================================================
ws2 = wb.create_sheet("Data Quality")
r = title_block(ws2, "Data Quality Assessment", "Missing values and outlier detection (application_data.csv)")

ws2.cell(row=r, column=1, value="Top 20 columns by % missing (application_data.csv, 307,511 rows)").font = BOLD_BODY
r += 1
miss = pd.read_csv('/home/claude/missing_pct.csv', index_col=0)
miss.columns = ['pct_missing']
miss.index.name = 'Column'
top_miss = miss.head(20)
r = write_df(ws2, top_miss, r, pct_cols=['pct_missing'])
r += 1

ws2.cell(row=r, column=1, value="Note: 49 columns exceeded 40% missing and were excluded from the core analysis (mostly building/apartment survey fields e.g. COMMONAREA_*, NONLIVINGAPARTMENTS_*, YEARS_BUILD_*).").font = SUBTITLE_FONT
r += 2

ws2.cell(row=r, column=1, value="Outlier scan — key numeric fields (IQR method, upper whisker = Q3 + 1.5xIQR)").font = BOLD_BODY
r += 1
outlier_data = pd.DataFrame({
    'Variable': ['AMT_INCOME_TOTAL','AMT_CREDIT','AMT_ANNUITY','AMT_GOODS_PRICE','CNT_CHILDREN','YEARS_EMPLOYED'],
    'Median': [147150, 513531, 24903, 450000, 0, 4.5],
    'Upper_Whisker': [337500, 1616625, 61704, 1341000, 2, 19],
    'Count_Above_Whisker': [14035, 6562, 7504, 14728, 4272, 15223],
    'Pct_Above_Whisker': [4.56, 2.13, 2.44, 4.79, 1.39, 4.95],
    'Max_Value': [117000000, 4050000, 258025.5, 4050000, 19, 49.1],
}).set_index('Variable')
r = write_df(ws2, outlier_data, r, pct_cols=['Pct_Above_Whisker'])
r += 1
ws2.cell(row=r, column=1, value="Note: AMT_INCOME_TOTAL has an extreme max (117M vs median 147K) — a small number of very high earners. CNT_CHILDREN max of 19 is implausible for most records and worth a manual data-integrity check. These outliers were retained (not removed) as they may reflect real high-net-worth applicants, but should be capped or reviewed before use in a scoring model.").font = SUBTITLE_FONT
ws2.row_dimensions[r].height = 30

autosize(ws2, 6)

# ============================================================
# SHEET 3: Target Imbalance
# ============================================================
ws3 = wb.create_sheet("Target Imbalance")
r = title_block(ws3, "TARGET Class Imbalance", "0 = repaid on time | 1 = payment difficulty")

target_df = pd.DataFrame({'Count':[282686,24825], 'Pct_of_Total':[91.93, 8.07]}, index=pd.Index(['0 - No difficulty','1 - Payment difficulty'], name='TARGET'))
r = write_df(ws3, target_df, r, pct_cols=['Pct_of_Total'])
data_start = r - 2  # header row + 1
chart = BarChart()
chart.title = "TARGET Distribution (%)"
chart.y_axis.title = "% of applicants"
chart.style = 10
cats = Reference(ws3, min_col=1, min_row=r-2, max_row=r-1)
data = Reference(ws3, min_col=3, min_row=r-3, max_row=r-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 14, 8
ws3.add_chart(chart, f"E{r-4}")
r += 12
ws3.cell(row=r, column=1, value="Implication: with ~12:1 imbalance, a naive model predicting 'no difficulty' for every applicant would already be 91.9% accurate — accuracy alone is a misleading metric here. Business rules and models must be evaluated on recall/precision for the minority (defaulter) class, not overall accuracy.").font = SUBTITLE_FONT
ws3.row_dimensions[r].height = 30
autosize(ws3, 3)

# ============================================================
# SHEET 4: Numeric Variables by Target (segmented univariate)
# ============================================================
ws4 = wb.create_sheet("Numeric Vars by Target")
r = title_block(ws4, "Key Numeric Variables — Median by TARGET", "Segmented univariate comparison")

seg = pd.read_csv('/home/claude/segment_medians.csv', index_col=0)
seg.columns = ['Target_0_Median','Target_1_Median']
seg.index.name='Variable'
seg['Pct_Difference'] = ((seg['Target_1_Median']-seg['Target_0_Median'])/seg['Target_0_Median']*100).round(1)
r = write_df(ws4, seg, r, pct_cols=['Pct_Difference'])
r += 2

ws4.cell(row=r, column=1, value="Default rate (%) by income bucket").font = BOLD_BODY
r += 1
income_bucket = pd.DataFrame({
    'Income_Bucket':['<100K','100-150K','150-200K','200-300K','300K+'],
    'Default_Rate_Pct':[8.20,8.62,8.45,7.55,5.95],
    'Applicant_Count':[63698,91591,64307,65176,22739]
}).set_index('Income_Bucket')
tbl_start = r
r = write_df(ws4, income_bucket, r, pct_cols=['Default_Rate_Pct'])
chart2 = BarChart()
chart2.title = "Default Rate by Income Bucket"
chart2.y_axis.title = "Default rate (%)"
cats2 = Reference(ws4, min_col=1, min_row=tbl_start+1, max_row=r-1)
data2 = Reference(ws4, min_col=2, min_row=tbl_start, max_row=r-1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width, chart2.height = 14, 8
ws4.add_chart(chart2, f"E{tbl_start}")
r += 12

ws4.cell(row=r, column=1, value="Default rate (%) by age bucket").font = BOLD_BODY
r += 1
age_bucket = pd.DataFrame({
    'Age_Bucket':['20-30','30-40','40-50','50-60','60-70'],
    'Default_Rate_Pct':[11.43,9.59,7.63,6.12,4.93],
    'Applicant_Count':[45389,82315,76504,68067,35236]
}).set_index('Age_Bucket')
tbl_start2 = r
r = write_df(ws4, age_bucket, r, pct_cols=['Default_Rate_Pct'])
chart3 = BarChart()
chart3.title = "Default Rate by Age Bucket"
chart3.y_axis.title = "Default rate (%)"
cats3 = Reference(ws4, min_col=1, min_row=tbl_start2+1, max_row=r-1)
data3 = Reference(ws4, min_col=2, min_row=tbl_start2, max_row=r-1)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
chart3.width, chart3.height = 14, 8
ws4.add_chart(chart3, f"E{tbl_start2}")
r += 12
autosize(ws4, 3)

# ============================================================
# SHEET 5: Categorical Variables by Target
# ============================================================
ws5 = wb.create_sheet("Categorical Vars by Target")
r = title_block(ws5, "Default Rate (%) by Categorical Segment", "Categories with fewer than 200 applicants excluded for statistical reliability")

cat_tables = {
 'Gender': pd.DataFrame({'Category':['M','F'],'Default_Rate_Pct':[10.14,7.00],'Count':[105059,202448]}),
 'Contract Type': pd.DataFrame({'Category':['Cash loans','Revolving loans'],'Default_Rate_Pct':[8.35,5.48],'Count':[278232,29279]}),
 'Income Type': pd.DataFrame({'Category':['Working','Commercial associate','State servant','Pensioner'],'Default_Rate_Pct':[9.59,7.48,5.75,5.39],'Count':[158774,71617,21703,55362]}),
 'Education': pd.DataFrame({'Category':['Lower secondary','Secondary/secondary special','Incomplete higher','Higher education'],'Default_Rate_Pct':[10.93,8.94,8.48,5.36],'Count':[3816,218391,10277,74863]}),
 'Family Status': pd.DataFrame({'Category':['Civil marriage','Single/not married','Separated','Married','Widow'],'Default_Rate_Pct':[9.94,9.81,8.19,7.56,5.82],'Count':[29775,45444,19770,196432,16088]}),
 'Housing Type': pd.DataFrame({'Category':['Rented apartment','With parents','Municipal apartment','Co-op apartment','House/apartment','Office apartment'],'Default_Rate_Pct':[12.31,11.70,8.54,7.93,7.80,6.57],'Count':[4881,14840,11183,1122,272868,2617]}),
 'Occupation (top 8 riskiest)': pd.DataFrame({'Category':['Low-skill Laborers','Drivers','Waiters/barmen staff','Security staff','Laborers','Cooking staff','Sales staff','Cleaning staff'],'Default_Rate_Pct':[17.15,11.33,11.28,10.74,10.58,10.44,9.63,9.61],'Count':[2093,18603,1348,6721,55186,5946,32102,4653]}),
 'Own Car': pd.DataFrame({'Category':['N','Y'],'Default_Rate_Pct':[8.50,7.24],'Count':[202924,104587]}),
 'Own Realty': pd.DataFrame({'Category':['N','Y'],'Default_Rate_Pct':[8.32,7.96],'Count':[94199,213312]}),
}

for name, tdf in cat_tables.items():
    tdf = tdf.set_index('Category')
    ws5.cell(row=r, column=1, value=name).font = BOLD_BODY
    r += 1
    tbl_start = r
    r = write_df(ws5, tdf, r, pct_cols=['Default_Rate_Pct'])
    if len(tdf) <= 8:
        chart = BarChart()
        chart.title = f"Default Rate by {name}"
        chart.y_axis.title = "%"
        cats = Reference(ws5, min_col=1, min_row=tbl_start+1, max_row=r-1)
        data = Reference(ws5, min_col=2, min_row=tbl_start, max_row=r-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 12, 7
        ws5.add_chart(chart, f"E{tbl_start}")
        r = max(r, tbl_start + 15)
    r += 2

autosize(ws5, 3)

# ============================================================
# SHEET 6: Correlation Analysis (by segment)
# ============================================================
ws6 = wb.create_sheet("Correlation Analysis")
r = title_block(ws6, "Correlation Matrix by TARGET Segment", "Numeric variables, computed separately for repayers (0) and defaulters (1)")

def write_corr_matrix(ws, corr_df, start_row, title):
    ws.cell(row=start_row, column=1, value=title).font = BOLD_BODY
    start_row += 1
    corr_round = corr_df.round(2)
    corr_round.index.name = 'Variable'
    rr = write_df(ws, corr_round, start_row, pct_cols=[])
    # conditional-ish manual highlight for |corr|>0.5 (excluding diagonal)
    for i, idx in enumerate(corr_round.index):
        for j, col in enumerate(corr_round.columns):
            if idx == col:
                continue
            val = corr_round.iloc[i, j]
            if abs(val) >= 0.5:
                cell = ws.cell(row=start_row+1+i, column=2+j)
                cell.fill = FLAG_FILL
    return rr

c0 = pd.read_csv('/home/claude/corr_target0.csv', index_col=0)
c1 = pd.read_csv('/home/claude/corr_target1.csv', index_col=0)
r = write_corr_matrix(ws6, c0, r, "Correlation matrix — TARGET = 0 (repaid on time)")
r += 2
r = write_corr_matrix(ws6, c1, r, "Correlation matrix — TARGET = 1 (payment difficulty)")
r += 2

ws6.cell(row=r, column=1, value="Largest shifts in correlation between segments (top 10)").font = BOLD_BODY
r += 1
diff = pd.read_csv('/home/claude/corr_diff.csv').head(10)
diff = diff.set_index(diff['var1'] + ' vs ' + diff['var2'])[['abs_diff','corr_target0','corr_target1']]
diff.columns=['Abs_Difference','Corr_Target0','Corr_Target1']
diff.index.name='Variable Pair'
r = write_df(ws6, diff.round(3), r)
r += 1
ws6.cell(row=r, column=1, value="Highlighted cells: |correlation| >= 0.5. Key finding: AMT_INCOME_TOTAL's correlation with AMT_ANNUITY/AMT_CREDIT/AMT_GOODS_PRICE is much weaker among defaulters (~0.04-0.05) than repayers (~0.35-0.42) — among defaulters, loan sizing is less anchored to actual income, a useful underwriting red flag.").font = SUBTITLE_FONT
ws6.row_dimensions[r].height = 30

autosize(ws6, 17, min_width=8, max_width=16)

# ============================================================
# SHEET 7: Previous Application Analysis
# ============================================================
ws7 = wb.create_sheet("Previous Application Analysis")
r = title_block(ws7, "Prior Loan History vs Current Default Risk", "previous_application.csv (1,670,214 records) merged to application_data.csv on SK_ID_CURR")

ws7.cell(row=r, column=1, value="Distribution of previous application decisions (all prior applications, all clients)").font = BOLD_BODY
r += 1
status_dist = pd.DataFrame({
    'Decision':['Approved','Canceled','Refused','Unused offer'],
    'Count':[1036781,316319,290678,26436],
    'Pct_of_Prior_Apps':[62.07,18.94,17.40,1.58]
}).set_index('Decision')
tbl_start = r
r = write_df(ws7, status_dist, r, pct_cols=['Pct_of_Prior_Apps'])
chart = BarChart()
chart.title = "Prior Application Decision Distribution"
chart.y_axis.title = "% of prior applications"
cats = Reference(ws7, min_col=1, min_row=tbl_start+1, max_row=r-1)
data = Reference(ws7, min_col=3, min_row=tbl_start, max_row=r-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 13, 7
ws7.add_chart(chart, f"E{tbl_start}")
r += 11

ws7.cell(row=r, column=1, value="Current default rate by prior refusal / cancellation history").font = BOLD_BODY
r += 1
hist_df = pd.DataFrame({
    'Client_History':['No prior refusal','Had >=1 prior refusal','No prior cancellation','Had >=1 prior cancellation'],
    'Current_Default_Rate_Pct':[6.98,10.32,7.75,8.65],
    'Applicant_Count':[207217,100294,196785,110726]
}).set_index('Client_History')
r = write_df(ws7, hist_df, r, pct_cols=['Current_Default_Rate_Pct'])
r += 2

ws7.cell(row=r, column=1, value="Current default rate by NUMBER of prior refusals (dose-response relationship)").font = BOLD_BODY
r += 1
refusal_dose = pd.DataFrame({
    'Num_Prior_Refusals':['0','1','2','3','4','5+'],
    'Current_Default_Rate_Pct':[6.98,8.83,10.24,11.43,12.02,14.09],
    'Applicant_Count':[207217,46534,22739,11859,6904,12258]
}).set_index('Num_Prior_Refusals')
tbl_start2 = r
r = write_df(ws7, refusal_dose, r, pct_cols=['Current_Default_Rate_Pct'])
chart2 = LineChart()
chart2.title = "Default Rate vs Number of Prior Refusals"
chart2.y_axis.title = "Default rate (%)"
chart2.x_axis.title = "Number of prior refusals"
cats2 = Reference(ws7, min_col=1, min_row=tbl_start2+1, max_row=r-1)
data2 = Reference(ws7, min_col=2, min_row=tbl_start2, max_row=r-1)
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.width, chart2.height = 14, 8
ws7.add_chart(chart2, f"E{tbl_start2}")
r += 12

ws7.cell(row=r, column=1, value="Key finding: this is a clean, near-monotonic dose-response relationship — each additional prior refusal from this lender raises the current default rate, from 6.98% (zero refusals) to 14.09% (5+ refusals), roughly a 2x increase. This is one of the strongest and most actionable signals in the dataset: repeated refusal history should escalate underwriting scrutiny or pricing regardless of the current application's other attributes.").font = SUBTITLE_FONT
ws7.row_dimensions[r].height = 30

autosize(ws7, 3)

# ============================================================
# SHEET 8: Recommendations
# ============================================================
ws8 = wb.create_sheet("Recommendations")
r = title_block(ws8, "Business Recommendations", "Translating EDA findings into underwriting actions")

recs = [
 ("1. Prioritize external bureau scores","EXT_SOURCE_2 and EXT_SOURCE_3 show the clearest separation between segments (median 0.57/0.55 for repayers vs 0.44/0.38 for defaulters). These should carry high weight in any scoring model or manual review checklist."),
 ("2. Use prior refusal count as an escalation trigger","Default rate rises from 6.98% (0 refusals) to 14.09% (5+ refusals) in a near-monotonic dose-response pattern. Recommend: 0 refusals = standard process; 1-2 = standard with note; 3+ = manual review or risk-based pricing."),
 ("3. Weight applicant age and employment tenure","Default rate falls from 11.4% (20-30 yrs) to 4.9% (60-70 yrs), and shorter YEARS_EMPLOYED correlates with higher default. Younger applicants with short tenure are not automatically declined but warrant additional income verification."),
 ("4. Flag high-risk occupation/income-type combinations","Low-skill Laborers (17.2%), Drivers (11.3%), and 'Working' income type (9.6%) default well above Pensioners (5.4%) or State servants (5.8%). Consider differentiated pricing or added conditions (co-signer, smaller loan-to-income ratio) for these segments rather than blanket rejection."),
 ("5. Do not rely on credit-to-income ratio alone","Median CREDIT_INCOME_RATIO is nearly identical between defaulters (3.25) and repayers (3.27) — it is a weak standalone signal. Pair it with EXT_SOURCE scores and history rather than using it as a primary cutoff."),
 ("6. Treat housing situation as a moderate signal","Renters and applicants living with parents default at 12.3% / 11.7% vs 7.8% for homeowners. Useful as a minor scoring factor, not a disqualifier on its own."),
 ("7. Recheck data quality before deploying a model","CNT_CHILDREN has implausible values (max 19) and AMT_INCOME_TOTAL has extreme outliers (max ~117M vs median ~147K) — cap or winsorize these before feeding into any predictive model to avoid distorted coefficients."),
]

for title_, text in recs:
    ws8.cell(row=r, column=1, value=title_).font = BOLD_BODY
    cell = ws8.cell(row=r, column=2, value=text)
    cell.font = BODY_FONT
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws8.row_dimensions[r].height = max(28, 14*(len(text)//85+1))
    r += 2

ws8.column_dimensions['A'].width = 32
ws8.column_dimensions['B'].width = 105

wb.save('/home/claude/credit_eda_case_study.xlsx')
print("Saved sheets 7-8")
